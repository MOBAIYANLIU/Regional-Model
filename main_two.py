
# 文件名：main_two.py
from CES_SIMULATOR_three import simulate_heating_system # 单building函数
from U_Value_Calculation import UVC # U-Value计算函数
import pandas as pd
import openpyxl
import os

def main():
    
    # 绝对路径设置
    base_dir = os.path.dirname(os.path.abspath(__file__)) # 当前脚本所在文件夹
    grid_path = os.path.join(base_dir, "grid emission factor.xlsx") # grid emission文件夹绝对路径
    input_path = os.path.join(base_dir, "street_level_three.xlsx") # data文件夹绝对路径
    
    grid_df = pd.read_excel(grid_path) # grid emission数据读取
    df = pd.read_excel(input_path) # 输入数据读取

    # Heater types
    heater_types = ["Gas", "ASHP", "DEH"]

    # 遍历每一年
    for index, grid_row in grid_df.iterrows():
        year = grid_row["Year"]
        emission_factor = grid_row["Emission Factor (kgCO2e/kWh)"]

        # 打开 street_level 为写入模式
        wb = openpyxl.load_workbook(input_path)
        ws = wb.active

        # 确定新列的索引
        base_col = ws.max_column + 1
        for i, ht in enumerate(heater_types):
            ws.cell(row=1, column=base_col + i * 2, value=f"{ht}_Heating_NPC_{year}")
            ws.cell(row=1, column=base_col + i * 2 + 1, value=f"{ht}_Heating_Emissions_{year}")

        # 遍历每个 property
        for i, row in df.iterrows():
            try:
                Location_Input = row["Postcode"]
                House_Size = row["House size"]
                Occupants_Num = row["Number habitable rooms"]
                EPC_Space_Heating = row["Energy consumption total"]
                TES_Max_Volume = 0.1
                Tstat = 20.0
                # 对于U_Value的设置，可以通过两个方式，一个是直接给数值（这个在初期仿真阶段测试代码用）；另一个是
                # 通过UVC函数来计算，这个主函数中提供了这两个方式，可以随时切换
                Dwelling_U_Value = UVC(House_Size,Location_Input,Occupants_Num,EPC_Space_Heating) # 通过函数
                # Dwelling_U_Value = 1.5 # 直接赋值
                Fixed_Grid_Emissions = emission_factor

                # 调用模拟函数
                Output_Record = simulate_heating_system(
                    Location_Input, House_Size, TES_Max_Volume,
                    Occupants_Num, Tstat, Dwelling_U_Value,
                    EPC_Space_Heating, Fixed_Grid_Emissions
                )

                # 初始化结果字典
                result_dict = {ht: {"NPC": None, "Emissions": None} for ht in heater_types}

                # 筛选 Eco7 的记录
                for record in Output_Record:
                    if "Eco7" in record:
                        for ht in heater_types:
                            if ht in record:
                                try:
                                    npc = record.split("NPC £")[1].split(",")[0].strip()
                                    ems = record.split("Emissions Heating")[1].split("kgCO2e")[0].strip()
                                    result_dict[ht]["NPC"] = npc
                                    result_dict[ht]["Emissions"] = ems
                                except:
                                    continue

                # 打印结果
                print(f"\nYear: {year}, Property ID: {i}")
                print(f"{'Heater':<10} {'Tariff':<10} {'NPC':<10} {'Emissions':<10}")
                for ht in heater_types:
                    npc = result_dict[ht]["NPC"]
                    ems = result_dict[ht]["Emissions"]
                    if npc and ems:
                        print(f"{ht:<10} {'Eco7':<10} {npc:<10} {ems:<10}")
                        ws.cell(row=i+2, column=base_col + heater_types.index(ht) * 2, value=npc)
                        ws.cell(row=i+2, column=base_col + heater_types.index(ht) * 2 + 1, value=ems)

            except Exception as e:
                print(f"Error in Property ID {i}, Year {year}: {e}")
                continue

        # 保存每年结果
        output_path = f"street_level_three_result_{year}.xlsx"
        wb.save(output_path)
        print(f"\nSimulation results for Year {year} saved to {output_path}")

if __name__ == "__main__":
    print("Hello, welcome to the Domestic Energy Simulator")
    main()
    print("The simulation is now complete, goodbye")
